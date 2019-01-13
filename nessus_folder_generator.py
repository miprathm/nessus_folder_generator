"""
1.Read xslx file
2.Split into two parts
3. path and filename with extension
4. Create folder of that file 
5. Read xlsx file 
"""

import os
import re
import openpyxl
import sys

pathname = os.path.abspath(sys.argv[0])
if len(sys.argv) > 1 :
	pathname = os.path.abspath(sys.argv[1])
print(pathname)

path, file = os.path.split(pathname) 

os.chdir(path)

filename, file_extension = os.path.splitext(file) 

#print(filename)
# Project directory
if not os.path.exists(filename):
    os.makedirs(filename)

wb_obj = openpyxl.load_workbook(pathname) 
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row = 1, column = 1) 
print(cell_obj.value)
print(sheet_obj.max_row)
temp = ''

for departmentIndex in range(1,sheet_obj.max_row):
	cell_obj = sheet_obj.cell(row = departmentIndex, column = 1) 
	print(cell_obj.value)



	



